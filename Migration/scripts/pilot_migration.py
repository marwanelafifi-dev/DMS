#!/usr/bin/env python3
"""Prepare and validate the first five-document KnowledgeTree migration pilot.

This script is intentionally DRY-RUN ONLY. It reads the legacy export, the two
owner-review workbooks, physical source blobs, and reference rows from the
local New DMS PostgreSQL database. It writes CSV/Markdown planning artifacts
under migration/output/ and never writes to PostgreSQL, MinIO, or the source
export.

Owner resolution is value-based. The original Author comes from
11_owner_author_classification.xlsx; exact or canonical/normalized Author
values are matched to Final DMS Mapping.xlsx. A normalized value with more
than one approved owner is quarantined as DOCUMENT_LEVEL_REVIEW_REQUIRED.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import subprocess
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from preflight_migration import (
    DOCUMENTS_COLS,
    DOCUMENT_CONTENT_VERSION_COLS,
    DOCUMENT_METADATA_VERSION_COLS,
    DOCUMENT_TYPES_LOOKUP_COLS,
    extract_table_rows,
    load_dump_text,
    rows_to_dicts,
)


SCRIPT_DIR = Path(__file__).resolve().parent
MIGRATION_DIR = SCRIPT_DIR.parent
REPO_DIR = MIGRATION_DIR.parent
SOURCE_DIR = MIGRATION_DIR / "source"
OUTPUT_DIR = MIGRATION_DIR / "output"

SOURCE_DUMP = SOURCE_DIR / "dms_full_2026-07-30.sql.gz"
BLOBS_DIR = SOURCE_DIR / "blobs"
METADATA_TSV = OUTPUT_DIR / "05_metadata_fields.tsv"
ORIGINAL_MAPPING_XLSX = OUTPUT_DIR / "11_owner_author_classification.xlsx"
FINAL_MAPPING_XLSX = OUTPUT_DIR / "Final DMS Mapping.xlsx"
MISMATCH_CSV = OUTPUT_DIR / "08_latest_file_mismatch_review.csv"

OWNER_RESOLUTION_CSV = OUTPUT_DIR / "12_owner_mapping_resolution.csv"
PILOT_PLAN_CSV = OUTPUT_DIR / "12_pilot_plan.csv"
DRY_RUN_REPORT_MD = OUTPUT_DIR / "12_pilot_dry_run_report.md"

EXCLUDED_DOCUMENT_IDS = {"164", "507", "928"}
EXPECTED_DOCUMENT_COUNT = 1008
EXPECTED_AUTHOR_VALUE_COUNT = 238
PILOT_CASES = (
    "Simple single legacy author",
    "Multiple legacy authors mapped to one approved new Owner",
    "Spelling/name-normalized legacy author",
    "Organization/non-person mapped to a real Owner",
    "Legacy author reassigned to a different current Owner",
)

# Corrections that occur inside multi-author strings but do not have their own
# canonical row in the classification workbook. These are spelling cleanup,
# not owner decisions. Owner decisions always come from Final DMS Mapping.xlsx.
KNOWN_COMPONENT_ALIASES = {
    "nabilelsenousy": "nabilsinoussi",
    "nabilelsinousi": "nabilsinoussi",
    "nabilsenousi": "nabilsinoussi",
    "ahmedsheble": "ahmedshebl",
}

KNOWN_FULL_VALUE_ALIASES = {
    # The original workbook classifies "Alaa" as an incomplete name with
    # "Alaa Fathy (possible only)"; the corrected final workbook uses Alaa
    # Fathy. The user explicitly requires that normalized value to be treated
    # as a split/document-level-review group.
    "alaa": "alaafathy",
}


@dataclass(frozen=True)
class OriginalAuthor:
    legacy_author: str
    document_count: int
    classification: str
    canonical_name: str


@dataclass(frozen=True)
class OwnerResolution:
    original: OriginalAuthor
    normalized_key: str
    corrected_values: tuple[str, ...]
    owners: tuple[str, ...]
    status: str

    @property
    def owner(self) -> str:
        return self.owners[0] if self.status == "UNAMBIGUOUS" else ""


def text(value: object | None) -> str:
    return "" if value is None else str(value).strip()


def compact_author(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).casefold()
    value = re.sub(r"\b(?:dr|eng)\.?\b", " ", value)
    value = re.sub(r"\band\b", " ", value)
    return "".join(ch for ch in value if ch.isalnum())


def normalize_owner(value: str) -> str:
    return "Amr El Mosallamy" if value.strip() == "Amr Mossallamy" else value.strip()


def load_workbook_rows() -> tuple[list[OriginalAuthor], list[dict[str, str]]]:
    original_sheet = load_workbook(
        ORIGINAL_MAPPING_XLSX, read_only=True, data_only=True
    )["All Authors"]
    final_book = load_workbook(FINAL_MAPPING_XLSX, read_only=True, data_only=True)
    final_sheet = final_book[final_book.sheetnames[0]]

    originals = []
    for row in original_sheet.iter_rows(min_row=2, values_only=True):
        originals.append(
            OriginalAuthor(
                legacy_author=text(row[0]),
                document_count=int(row[1]),
                classification=text(row[3]),
                canonical_name=text(row[4]),
            )
        )

    finals = []
    for row in final_sheet.iter_rows(min_row=2, values_only=True):
        finals.append(
            {
                "corrected_author": text(row[0]),
                "approved_owner": normalize_owner(text(row[2])),
            }
        )
    return originals, finals


def build_aliases(originals: list[OriginalAuthor]) -> dict[str, str]:
    aliases = dict(KNOWN_COMPONENT_ALIASES)
    for original in originals:
        if not original.canonical_name:
            continue
        source = compact_author(original.legacy_author)
        target = compact_author(original.canonical_name)
        if source and source != target:
            aliases[source] = target
    return aliases


def normalized_author_key(value: str, aliases: dict[str, str]) -> str:
    key = compact_author(value)
    # Longest first prevents a shorter spelling alias from consuming part of a
    # longer name. Replacement is also what lets a canonical single-person
    # spelling normalize that same person inside a multiple-author value.
    for source, target in sorted(aliases.items(), key=lambda item: -len(item[0])):
        key = key.replace(source, target)
    return KNOWN_FULL_VALUE_ALIASES.get(key, key)


def resolve_owners() -> tuple[list[OwnerResolution], dict[str, OwnerResolution]]:
    originals, finals = load_workbook_rows()
    if len(originals) != EXPECTED_AUTHOR_VALUE_COUNT or len(finals) != EXPECTED_AUTHOR_VALUE_COUNT:
        raise RuntimeError(
            f"Expected 238 data rows in each owner workbook; found "
            f"{len(originals)} original and {len(finals)} final rows"
        )
    if len({row.legacy_author for row in originals}) != len(originals):
        raise RuntimeError("Original workbook Legacy Author values are not unique")

    aliases = build_aliases(originals)
    final_by_key: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in finals:
        key = normalized_author_key(row["corrected_author"], aliases)
        if not key or not row["approved_owner"]:
            raise RuntimeError(f"Incomplete final mapping row: {row}")
        final_by_key[key].append(row)

    resolutions = []
    by_original = {}
    for original in originals:
        source_value = original.canonical_name or original.legacy_author
        key = normalized_author_key(source_value, aliases)
        candidates = final_by_key.get(key, [])
        corrected = tuple(sorted({row["corrected_author"] for row in candidates}, key=str.casefold))
        owners = tuple(sorted({row["approved_owner"] for row in candidates}, key=str.casefold))
        if len(owners) == 1:
            status = "UNAMBIGUOUS"
        elif len(owners) > 1:
            status = "DOCUMENT_LEVEL_REVIEW_REQUIRED"
        else:
            status = "UNRESOLVED"
        resolution = OwnerResolution(original, key, corrected, owners, status)
        resolutions.append(resolution)
        by_original[original.legacy_author] = resolution

    known_split_keys = {
        normalized_author_key(value, aliases)
        for value in (
            "Abdelrahman El Naggar",
            "Ahmad Shereef",
            "Alaa Fathy",
            "Mahmoud Yousef",
        )
    }
    status_by_key = {row.normalized_key: row.status for row in resolutions}
    missed_known_splits = sorted(
        key for key in known_split_keys
        if status_by_key.get(key) != "DOCUMENT_LEVEL_REVIEW_REQUIRED"
    )
    if missed_known_splits:
        raise RuntimeError(
            "Known split-owner examples were not quarantined: " + ", ".join(missed_known_splits)
        )
    return resolutions, by_original


def write_owner_resolution(resolutions: list[OwnerResolution]) -> None:
    fields = [
        "legacy_author_original",
        "document_count",
        "classification",
        "canonical_normalized_name",
        "match_method",
        "normalized_author_key",
        "final_corrected_legacy_author_values",
        "approved_new_dms_owners",
        "resolution_status",
    ]
    with OWNER_RESOLUTION_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for resolution in resolutions:
            if resolution.original.legacy_author in resolution.corrected_values:
                match_method = "EXACT_VALUE"
            elif resolution.original.canonical_name:
                match_method = "CANONICAL_NORMALIZED_NAME"
            else:
                match_method = "NORMALIZED_VALUE"
            writer.writerow(
                {
                    "legacy_author_original": resolution.original.legacy_author,
                    "document_count": resolution.original.document_count,
                    "classification": resolution.original.classification,
                    "canonical_normalized_name": resolution.original.canonical_name,
                    "match_method": match_method,
                    "normalized_author_key": resolution.normalized_key,
                    "final_corrected_legacy_author_values": " | ".join(resolution.corrected_values),
                    "approved_new_dms_owners": " | ".join(resolution.owners),
                    "resolution_status": resolution.status,
                }
            )


def load_current_metadata() -> dict[str, dict[str, str]]:
    current: dict[str, dict[str, str]] = defaultdict(dict)
    with METADATA_TSV.open(encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            doc_id = row["doc_id"]
            if row["is_current_metadata_version"] == "1":
                current[doc_id][row["field_name"]] = row["field_value"]
    return dict(current)


def load_legacy_data() -> dict[str, object]:
    dump_text = load_dump_text(SOURCE_DUMP)
    documents = rows_to_dicts(extract_table_rows(dump_text, "documents"), DOCUMENTS_COLS)
    content_versions = rows_to_dicts(
        extract_table_rows(dump_text, "document_content_version"), DOCUMENT_CONTENT_VERSION_COLS
    )
    metadata_versions = rows_to_dicts(
        extract_table_rows(dump_text, "document_metadata_version"), DOCUMENT_METADATA_VERSION_COLS
    )
    document_types = rows_to_dicts(
        extract_table_rows(dump_text, "document_types_lookup"), DOCUMENT_TYPES_LOOKUP_COLS
    )
    return {
        "documents": documents,
        "document_by_id": {row["id"]: row for row in documents},
        "version_by_id": {row["id"]: row for row in content_versions},
        "versions_by_doc": group_by(content_versions, "document_id"),
        "metadata_by_id": {row["id"]: row for row in metadata_versions},
        "metadata_by_doc": group_by(metadata_versions, "document_id"),
        "document_type_by_id": {row["id"]: row["name"] for row in document_types},
    }


def group_by(rows: list[dict[str, str]], key: str) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row[key]].append(row)
    return dict(grouped)


def load_mismatch_document_ids() -> set[str]:
    with MISMATCH_CSV.open(encoding="utf-8-sig", newline="") as handle:
        return {row["legacy_doc_id"] for row in csv.DictReader(handle)}


def md5_file(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_active_file(version: dict[str, str]) -> bool:
    storage_path = text(version.get("storage_path"))
    expected_md5 = text(version.get("md5hash")).casefold()
    if not storage_path or not expected_md5:
        return False
    path = BLOBS_DIR / Path(storage_path)
    return path.is_file() and path.stat().st_size > 0 and md5_file(path).casefold() == expected_md5


def transformed_original_document_id(document_number: str) -> str:
    return "" if document_number.strip().casefold() == "external document" else document_number.strip()


def tags_from_legacy(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def pilot_case_for(resolution: OwnerResolution) -> set[str]:
    original = resolution.original
    classification = original.classification
    original_key = compact_author(original.legacy_author)
    canonical_key = compact_author(original.canonical_name or original.legacy_author)
    owner_key = compact_author(resolution.owner)
    cases = set()
    if classification == "Multiple Authors":
        cases.add(PILOT_CASES[1])
    if original.canonical_name and original_key != canonical_key:
        cases.add(PILOT_CASES[2])
    if "Organization / Non-person" in classification:
        cases.add(PILOT_CASES[3])
    is_single_person = classification.startswith("Person -")
    if is_single_person and original_key == owner_key and original_key == canonical_key:
        cases.add(PILOT_CASES[0])
    if is_single_person and canonical_key != owner_key:
        cases.add(PILOT_CASES[4])
    return cases


def build_document_candidates(
    legacy: dict[str, object],
    current_metadata: dict[str, dict[str, str]],
    resolutions: dict[str, OwnerResolution],
) -> tuple[dict[str, list[dict[str, object]]], dict[str, str]]:
    documents = legacy["documents"]
    metadata_by_id = legacy["metadata_by_id"]
    version_by_id = legacy["version_by_id"]
    versions_by_doc = legacy["versions_by_doc"]
    metadata_versions_by_doc = legacy["metadata_by_doc"]
    document_type_by_id = legacy["document_type_by_id"]
    mismatch_ids = load_mismatch_document_ids()
    excluded = EXCLUDED_DOCUMENT_IDS | mismatch_ids
    by_case: dict[str, list[dict[str, object]]] = defaultdict(list)
    document_mapping_status: dict[str, str] = {}

    for document in documents:
        doc_id = document["id"]
        fields = current_metadata.get(doc_id, {})
        author = fields.get("Authors", "")
        resolution = resolutions.get(author)
        if resolution is None:
            raise RuntimeError(f"Document {doc_id} has an Author absent from the original workbook: {author!r}")
        document_mapping_status[doc_id] = resolution.status
        if resolution.status != "UNAMBIGUOUS" or doc_id in excluded:
            continue

        metadata_version = metadata_by_id.get(document["metadata_version_id"])
        if metadata_version is None:
            continue
        active_version = version_by_id.get(metadata_version["content_version_id"])
        if active_version is None or active_version["document_id"] != doc_id:
            continue
        if not clean_active_file(active_version):
            continue

        document_number = fields.get("Document #", "")
        group = fields.get("Group", "")
        document_type = document_type_by_id.get(metadata_version["document_type_id"], "")
        corrected = " | ".join(resolution.corrected_values)
        row: dict[str, object] = {
            "legacy_document_id": doc_id,
            "title": metadata_version["name"] or "",
            "legacy_author_original": author,
            "legacy_author_corrected": corrected,
            "new_owner_name": resolution.owner,
            "group": group,
            "new_department": group,
            "document_type": document_type,
            "new_category": document_type,
            "document_number": document_number,
            "original_document_id": transformed_original_document_id(document_number),
            "description": fields.get("Description", ""),
            "tags": " | ".join(tags_from_legacy(fields.get("Tag", ""))),
            "active_content_version_id": active_version["id"],
            "active_filename": active_version["filename"] or "",
            "active_blob_path": f"migration/source/blobs/{active_version['storage_path']}",
            "active_md5": active_version["md5hash"] or "",
            "historical_file_version_count": max(0, len(versions_by_doc.get(doc_id, [])) - 1),
            "metadata_history_count": max(0, len(metadata_versions_by_doc.get(doc_id, [])) - 1),
            "legacy_folder_path": (document.get("full_path") or "").strip("/"),
            "target_folder_path": (document.get("full_path") or "").strip("/"),
            "owner_mapping_status": resolution.status,
        }
        for case in pilot_case_for(resolution):
            candidate = dict(row)
            candidate["pilot_case"] = case
            by_case[case].append(candidate)

    for case in by_case:
        # Prefer rows that exercise archive linking, then stable smallest ID.
        by_case[case].sort(
            key=lambda row: (
                -int(row["historical_file_version_count"] > 0),
                -int(row["metadata_history_count"] > 0),
                int(row["legacy_document_id"]),
            )
        )
    return dict(by_case), document_mapping_status


def select_pilot(by_case: dict[str, list[dict[str, object]]]) -> list[dict[str, object]]:
    selected = []
    used_ids = set()
    for case in PILOT_CASES:
        candidate = next(
            (row for row in by_case.get(case, []) if row["legacy_document_id"] not in used_ids),
            None,
        )
        if candidate is None:
            raise RuntimeError(f"No eligible clean, unambiguous document found for pilot case: {case}")
        selected.append(candidate)
        used_ids.add(candidate["legacy_document_id"])
    if len(selected) != 5 or len(used_ids) != 5:
        raise RuntimeError("Pilot selection did not produce exactly five distinct documents")
    return selected


def write_pilot_plan(selected: list[dict[str, object]]) -> None:
    fields = [
        "legacy_document_id", "title", "legacy_author_original",
        "legacy_author_corrected", "new_owner_name", "group", "new_department",
        "document_type", "new_category", "document_number", "original_document_id",
        "description", "tags", "active_content_version_id", "active_filename",
        "active_blob_path", "active_md5", "historical_file_version_count",
        "metadata_history_count", "legacy_folder_path", "target_folder_path",
        "owner_mapping_status", "pilot_case",
    ]
    with PILOT_PLAN_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(selected)


def parse_env_file() -> dict[str, str]:
    result = {}
    for raw_line in (REPO_DIR / ".env").read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        result[key.strip()] = value.strip().strip('"').strip("'")
    return result


def psql_rows(query: str) -> list[list[str]]:
    env = parse_env_file()
    command = [
        "docker", "compose", "exec", "-T", "postgres", "psql",
        "-U", env["POSTGRES_USER"], "-d", env["POSTGRES_DB"],
        "-v", "ON_ERROR_STOP=1", "-At", "-F", "\t", "-c", query,
    ]
    completed = subprocess.run(
        command,
        cwd=REPO_DIR,
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    return [line.split("\t") for line in completed.stdout.splitlines() if line]


def load_new_dms_state() -> dict[str, object]:
    users = {row[0].casefold(): row[0] for row in psql_rows(
        "SELECT full_name FROM dms_users WHERE is_active = TRUE ORDER BY lower(full_name);"
    )}
    dropdowns: dict[str, set[str]] = defaultdict(set)
    for list_key, label in psql_rows(
        "SELECT list_key, label FROM dms_dropdown_items "
        "WHERE list_key IN ('department','category') ORDER BY list_key, label;"
    ):
        dropdowns[list_key].add(label.casefold())
    folders = {row[0].casefold(): row[0] for row in psql_rows(
        "WITH RECURSIVE tree AS ("
        "SELECT folder_id,parent_folder_id,name,name::text AS path FROM dms_folders WHERE parent_folder_id IS NULL "
        "UNION ALL SELECT f.folder_id,f.parent_folder_id,f.name,(t.path || '/' || f.name)::text "
        "FROM dms_folders f JOIN tree t ON t.folder_id=f.parent_folder_id) "
        "SELECT path FROM tree ORDER BY path;"
    )}
    documents = psql_rows("SELECT COALESCE(original_document_id,'') FROM dms_documents;")
    return {
        "users": users,
        "departments": dropdowns["department"],
        "categories": dropdowns["category"],
        "folders": folders,
        "original_document_ids": {row[0].casefold() for row in documents if row[0]},
    }


def validate_selected(
    selected: list[dict[str, object]],
    legacy: dict[str, object],
    current_metadata: dict[str, dict[str, str]],
    new_dms: dict[str, object],
) -> tuple[list[dict[str, object]], bool]:
    document_by_id = legacy["document_by_id"]
    metadata_by_id = legacy["metadata_by_id"]
    metadata_by_doc = legacy["metadata_by_doc"]
    version_by_id = legacy["version_by_id"]
    versions_by_doc = legacy["versions_by_doc"]
    document_type_by_id = legacy["document_type_by_id"]
    planned_original_ids = Counter(
        str(row["original_document_id"]).casefold()
        for row in selected if str(row["original_document_id"])
    )
    results = []
    for row in selected:
        blob = REPO_DIR / str(row["active_blob_path"])
        actual_md5 = md5_file(blob) if blob.is_file() else ""
        source_ok = blob.is_file() and blob.stat().st_size > 0
        hash_ok = source_ok and actual_md5.casefold() == str(row["active_md5"]).casefold()
        owner_ok = str(row["new_owner_name"]).casefold() in new_dms["users"]
        department_ok = str(row["new_department"]).casefold() in new_dms["departments"]
        category_ok = str(row["new_category"]).casefold() in new_dms["categories"]
        folder_ok = str(row["target_folder_path"]).casefold() in new_dms["folders"]
        doc_id = str(row["legacy_document_id"])
        document = document_by_id[doc_id]
        active_metadata = metadata_by_id[document["metadata_version_id"]]
        fields = current_metadata[doc_id]
        expected_type = document_type_by_id[active_metadata["document_type_id"]]
        metadata_ok = (
            str(row["legacy_author_original"]) == fields.get("Authors", "")
            and str(row["group"]) == fields.get("Group", "")
            and str(row["new_department"]) == fields.get("Group", "")
            and str(row["document_type"]) == expected_type
            and str(row["new_category"]) == expected_type
            and str(row["description"]) == fields.get("Description", "")
            and str(row["tags"]) == " | ".join(tags_from_legacy(fields.get("Tag", "")))
            and str(row["document_number"]) == fields.get("Document #", "")
            and str(row["original_document_id"])
            == transformed_original_document_id(fields.get("Document #", ""))
            and str(row["title"]) == text(active_metadata["name"])
            and str(row["active_content_version_id"]) == active_metadata["content_version_id"]
        )
        metadata_versions = metadata_by_doc.get(doc_id, [])
        content_versions = versions_by_doc.get(doc_id, [])
        metadata_content_links_ok = all(
            not item["content_version_id"]
            or (
                item["content_version_id"] in version_by_id
                and version_by_id[item["content_version_id"]]["document_id"] == doc_id
            )
            for item in metadata_versions
        )
        archive_ok = (
            bool(metadata_versions)
            and bool(content_versions)
            and active_metadata in metadata_versions
            and metadata_content_links_ok
            and all(item["document_id"] == doc_id and item["id"] for item in content_versions)
        )
        original_id = str(row["original_document_id"]).casefold()
        duplicate_original_id = bool(original_id and original_id in new_dms["original_document_ids"])
        duplicate_within_plan = bool(original_id and planned_original_ids[original_id] > 1)
        duplicate_ok = not duplicate_original_id and not duplicate_within_plan
        checks = {
            "source_file_exists_nonzero": source_ok,
            "md5_matches": hash_ok,
            "target_owner_exists": owner_ok,
            "department_valid": department_ok,
            "category_valid": category_ok,
            "folder_mapping_resolved": folder_ok,
            "metadata_transformable": metadata_ok,
            "archive_history_linkable": archive_ok,
            "no_duplicate_target_record": duplicate_ok,
        }
        results.append({"plan": row, "checks": checks, "ready": all(checks.values())})
    return results, all(result["ready"] for result in results)


def write_report(
    resolutions: list[OwnerResolution],
    document_mapping_status: dict[str, str],
    results: list[dict[str, object]],
    ready: bool,
) -> None:
    unambiguous = sum(row.status == "UNAMBIGUOUS" for row in resolutions)
    review = sum(row.status == "DOCUMENT_LEVEL_REVIEW_REQUIRED" for row in resolutions)
    unresolved = sum(row.status == "UNRESOLVED" for row in resolutions)
    covered = len(document_mapping_status)
    unambiguous_docs = sum(status == "UNAMBIGUOUS" for status in document_mapping_status.values())
    review_docs = sum(
        status == "DOCUMENT_LEVEL_REVIEW_REQUIRED" for status in document_mapping_status.values()
    )
    lines = [
        "# First KnowledgeTree Migration Pilot - Dry-Run Report\n",
        "\n",
        "DRY-RUN only. The script performed read-only source-file checks and read-only PostgreSQL "
        "reference queries. It did not write to PostgreSQL or MinIO and did not migrate any record.\n",
        "\n## Owner mapping resolution\n",
        f"\n- Original legacy Author values: **{len(resolutions)}**\n",
        f"- Resolved unambiguously: **{unambiguous}**\n",
        f"- Document-level review required: **{review}**\n",
        f"- Unresolved/no normalized match: **{unresolved}**\n",
        f"- Document-level mapping coverage: **{covered} / {EXPECTED_DOCUMENT_COUNT}**\n",
        f"- Documents with unambiguous mappings: **{unambiguous_docs}**\n",
        f"- Documents quarantined for document-level review: **{review_docs}**\n",
        "\nThe mapping is value-based; workbook row position is never used. Owner spelling "
        "`Amr Mossallamy` is normalized to `Amr El Mosallamy`.\n",
        "\n### Values requiring document-level review\n",
        "\n| Original legacy Author | Normalized value | Approved Owners |\n",
        "|---|---|---|\n",
    ]
    for resolution in resolutions:
        if resolution.status == "DOCUMENT_LEVEL_REVIEW_REQUIRED":
            lines.append(
                f"| {resolution.original.legacy_author} | "
                f"{' / '.join(resolution.corrected_values)} | "
                f"{' / '.join(resolution.owners)} |\n"
            )
    lines.extend([
        "\n## Selected pilot and validation\n",
        "\nFolder resolution uses the proposed preserved legacy path and passes only when that "
        "exact path already exists in the local New DMS; the dry-run never creates folders.\n",
        "\n| Legacy ID | Pilot case | Approved Owner | Active filename | Source | MD5 | Owner | Department | Category | Folder | Metadata | Archive links | Duplicate | Result |\n",
        "|---:|---|---|---|---|---|---|---|---|---|---|---|---|---|\n",
    ])
    for result in results:
        row = result["plan"]
        checks = result["checks"]
        mark = lambda value: "PASS" if value else "FAIL"
        lines.append(
            f"| {row['legacy_document_id']} | {row['pilot_case']} | {row['new_owner_name']} | "
            f"{row['active_filename']} | {mark(checks['source_file_exists_nonzero'])} | "
            f"{mark(checks['md5_matches'])} | {mark(checks['target_owner_exists'])} | "
            f"{mark(checks['department_valid'])} | {mark(checks['category_valid'])} | "
            f"{mark(checks['folder_mapping_resolved'])} | {mark(checks['metadata_transformable'])} | "
            f"{mark(checks['archive_history_linkable'])} | "
            f"{mark(checks['no_duplicate_target_record'])} | "
            f"{'READY' if result['ready'] else 'NOT READY'} |\n"
        )
    blocking = []
    for result in results:
        failed = [name for name, passed in result["checks"].items() if not passed]
        if failed:
            blocking.append(
                f"- Document {result['plan']['legacy_document_id']}: " + ", ".join(failed) + "\n"
            )
    lines.extend([
        "\n## Blocking issues\n",
        "\n" + ("".join(blocking) if blocking else "- None.\n"),
        f"\n## Final status: {'READY' if ready else 'NOT READY'}\n",
    ])
    DRY_RUN_REPORT_MD.write_text("".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dry-run", action="store_true", default=True,
        help="validate and write planning artifacts only (default and only supported mode)",
    )
    parser.add_argument(
        "--execute", action="store_true",
        help="refused: actual migration is intentionally not implemented for this pilot step",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.execute:
        print("ERROR: actual migration is disabled; run DRY-RUN only", file=sys.stderr)
        return 2
    required = [SOURCE_DUMP, BLOBS_DIR, METADATA_TSV, ORIGINAL_MAPPING_XLSX,
                FINAL_MAPPING_XLSX, MISMATCH_CSV, REPO_DIR / ".env"]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise RuntimeError("Missing required input(s): " + ", ".join(missing))

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    resolutions, resolutions_by_original = resolve_owners()
    write_owner_resolution(resolutions)
    unresolved = [row for row in resolutions if row.status == "UNRESOLVED"]
    if unresolved:
        raise RuntimeError(
            f"{len(unresolved)} original Author values have no normalized Final mapping; stopping"
        )

    current_metadata = load_current_metadata()
    legacy = load_legacy_data()
    if len(legacy["documents"]) != EXPECTED_DOCUMENT_COUNT:
        raise RuntimeError(
            f"Expected {EXPECTED_DOCUMENT_COUNT} documents; found {len(legacy['documents'])}"
        )
    by_case, document_mapping_status = build_document_candidates(
        legacy, current_metadata, resolutions_by_original
    )
    if len(document_mapping_status) != EXPECTED_DOCUMENT_COUNT:
        raise RuntimeError(
            f"Owner mapping covers {len(document_mapping_status)} of {EXPECTED_DOCUMENT_COUNT} documents"
        )
    selected = select_pilot(by_case)
    write_pilot_plan(selected)

    new_dms = load_new_dms_state()
    results, ready = validate_selected(selected, legacy, current_metadata, new_dms)
    write_report(resolutions, document_mapping_status, results, ready)

    print(f"Resolved unambiguously: {sum(r.status == 'UNAMBIGUOUS' for r in resolutions)}")
    print(
        "Document-level review required: "
        f"{sum(r.status == 'DOCUMENT_LEVEL_REVIEW_REQUIRED' for r in resolutions)}"
    )
    for row in selected:
        print(
            f"{row['legacy_document_id']} | {row['pilot_case']} | "
            f"{row['new_owner_name']} | {row['active_filename']}"
        )
    print("READY" if ready else "NOT READY")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
